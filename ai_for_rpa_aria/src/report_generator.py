"""
report_generator.py – Geração de Artefatos e Outputs Técnicos

Responsabilidades:
  • Exportar dados para CSV (csv nativo)
  • Gerar relatório JSON consolidado
  • Produzir relatório textual (TXT) com análise de IA
  • Criar planilha Excel (openpyxl) com múltiplas abas
"""

import csv
import json
import logging
from datetime import datetime, timezone
from pathlib import Path

log = logging.getLogger("ARIA.Reporter")


class ReportGenerator:
    """Gerador de artefatos técnicos do pipeline ARIA"""

    def __init__(self, db):
        self.db = db

    # ── CSV – Artigos ─────────────────────────────────────────────────────
    def export_articles_csv(self, filepath: str = "reports/articles.csv") -> str:
        """Exporta artigos de notícias para CSV."""
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        articles = self.db.get_articles(limit=100)

        if not articles:
            log.warning("Nenhum artigo para exportar.")
            return filepath

        fields = ["id", "title", "news_site", "published_at", "summary", "url"]

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(articles)

        log.info("✔ CSV de artigos exportado: %s (%d linhas)", filepath, len(articles))
        return filepath

    # ── CSV – NEOs ────────────────────────────────────────────────────────
    def export_neos_csv(self, filepath: str = "reports/neos.csv") -> str:
        """Exporta NEOs para CSV com flag de periculosidade."""
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        neos = self.db.get_neos()

        if not neos:
            log.warning("Nenhum NEO para exportar.")
            return filepath

        fields = [
            "neo_id", "name", "is_hazardous", "diameter_km_min",
            "diameter_km_max", "velocity_kph", "miss_distance_km",
            "approach_date", "nasa_url",
        ]

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()

            for neo in neos:
                neo["is_hazardous"] = "SIM" if neo.get("is_hazardous") else "NÃO"
                writer.writerow(neo)

        log.info("✔ CSV de NEOs exportado: %s (%d linhas)", filepath, len(neos))
        return filepath

    # ── JSON – Dados consolidados ─────────────────────────────────────────
    def export_full_json(self, filepath: str = "reports/mission_data.json") -> str:
        """Exporta todos os dados em um único JSON consolidado."""
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)

        payload = {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "generator":    "ARIA v1.0 – FIAP Global Solution 2026",
            "stats":        self.db.get_stats(),
            "articles":     self.db.get_articles(limit=50),
            "neos":         self.db.get_neos(),
            "iss_latest":   self.db.get_latest_iss(),
            "ai_analysis":  self.db.get_latest_analysis(),
        }

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        log.info("✔ JSON consolidado exportado: %s", filepath)
        return filepath

    # ── TXT – Relatório executivo ─────────────────────────────────────────
    def export_summary_txt(
        self,
        filepath:     str  = "reports/summary.txt",
        analysis:     dict | None = None,
        sys_snapshot: dict | None = None,
    ) -> str:
        """Gera relatório textual legível com todos os dados relevantes."""
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)

        stats    = self.db.get_stats()
        neos     = self.db.get_neos()
        iss      = self.db.get_latest_iss() or {}
        analysis = analysis or {}
        sys_snap = sys_snapshot or {}

        hazardous = [n for n in neos if n.get("is_hazardous") in (1, True, "SIM")]
        closest   = neos[0] if neos else {}

        lines = [
            "=" * 70,
            "  ARIA – Automated Reconnaissance & Intelligence for Astronomy",
            "  FIAP | Global Solution 2026 | AI for RPA",
            f"  Gerado em: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "=" * 70,
            "",
            "── ESTATÍSTICAS DA COLETA ─────────────────────────────────────────",
            f"  Artigos coletados : {stats.get('total_articles', 0)}",
            f"  NEOs monitorados  : {stats.get('total_neos', 0)}",
            f"  NEOs perigosos    : {stats.get('hazardous_neos', 0)}",
            f"  Snapshots ISS     : {stats.get('iss_snapshots', 0)}",
            f"  Análises de IA    : {stats.get('ai_analyses', 0)}",
            "",
            "── POSIÇÃO ATUAL DA ISS ───────────────────────────────────────────",
            f"  Latitude  : {iss.get('latitude', 'N/A')}°",
            f"  Longitude : {iss.get('longitude', 'N/A')}°",
            f"  Capturado : {iss.get('collected_at', 'N/A')}",
            "",
            "── OBJETO MAIS PRÓXIMO DA TERRA (NEO) ─────────────────────────────",
        ]

        if closest:
            lines += [
                f"  Nome        : {closest.get('name', 'N/A')}",
                f"  Distância   : {closest.get('miss_distance_km', 0):,.0f} km",
                f"  Velocidade  : {closest.get('velocity_kph', 0):,.0f} km/h",
                f"  Diâmetro    : {closest.get('diameter_km_min', 0):.3f} – "
                f"{closest.get('diameter_km_max', 0):.3f} km",
                f"  ⚠ Perigoso  : {'SIM' if closest.get('is_hazardous') else 'NÃO'}",
            ]
        else:
            lines.append("  Nenhum NEO disponível.")

        lines += [
            "",
            f"  Total de NEOs perigosos nas próximas 48h: {len(hazardous)}",
            "",
            "── MONITORAMENTO DO SISTEMA (psutil) ──────────────────────────────",
            f"  CPU        : {sys_snap.get('cpu_percent', 'N/A')}%",
            f"  RAM usada  : {sys_snap.get('ram_percent', 'N/A')}%",
            f"  Disco      : {sys_snap.get('disk_percent', 'N/A')}%",
            f"  PID ARIA   : {sys_snap.get('proc_pid', 'N/A')}",
            f"  Threads    : {sys_snap.get('proc_threads', 'N/A')}",
            "",
            "── ANÁLISE DE INTELIGÊNCIA ARTIFICIAL ──────────────────────────────",
            "",
            "  [RESUMO EXECUTIVO]",
            f"  {analysis.get('summary', 'N/A')}",
            "",
            "  [ALERTA DE AMEAÇAS]",
            f"  {analysis.get('hazard_alert', 'N/A')}",
            "",
            "  [INSIGHTS ESTRATÉGICOS]",
            f"  {analysis.get('insights', 'N/A')}",
            "",
            "=" * 70,
            "  Artefatos exportados:",
            "    • reports/articles.csv   – Notícias espaciais (CSV)",
            "    • reports/neos.csv       – Objetos próximos da Terra (CSV)",
            "    • reports/mission_data.json – Dados consolidados (JSON)",
            "    • reports/summary.txt    – Este relatório (TXT)",
            "=" * 70,
        ]

        with open(filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")

        log.info("✔ Relatório TXT exportado: %s", filepath)
        return filepath

    # ── Excel – Planilha completa (openpyxl) ──────────────────────────────
    def export_excel(self, filepath: str = "reports/aria_mission.xlsx") -> str:
        """
        Gera planilha Excel com 3 abas: Artigos, NEOs, Estatísticas.
        Demonstra: leitura/escrita de arquivos no formato XLSX.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            log.warning("openpyxl não instalado – Excel não gerado. Instale com: pip install openpyxl")
            return ""

        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()

        # ── Aba 1: Artigos ───────────────────────────────────────────────
        ws_articles = wb.active
        ws_articles.title = "Artigos"
        header_style = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", fgColor="1A1A2E")

        articles_headers = ["ID", "Título", "Site", "Publicado em", "Resumo", "URL"]
        for col, h in enumerate(articles_headers, 1):
            cell = ws_articles.cell(1, col, h)
            cell.font  = header_style
            cell.fill  = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row, art in enumerate(self.db.get_articles(50), 2):
            ws_articles.append([
                art.get("id"), art.get("title"), art.get("news_site"),
                art.get("published_at"), art.get("summary", "")[:200], art.get("url"),
            ])

        # ── Aba 2: NEOs ──────────────────────────────────────────────────
        ws_neos = wb.create_sheet("NEOs")
        neos_headers = ["ID", "Nome", "Perigoso?", "Diâm. Mín (km)", "Diâm. Máx (km)",
                        "Velocidade (km/h)", "Distância Terra (km)", "Data Aproximação"]
        for col, h in enumerate(neos_headers, 1):
            cell = ws_neos.cell(1, col, h)
            cell.font = header_style
            cell.fill = header_fill

        for row, neo in enumerate(self.db.get_neos(), 2):
            ws_neos.append([
                neo.get("id"), neo.get("name"),
                "⚠ SIM" if neo.get("is_hazardous") else "NÃO",
                neo.get("diameter_km_min"), neo.get("diameter_km_max"),
                neo.get("velocity_kph"), neo.get("miss_distance_km"), neo.get("approach_date"),
            ])

        # ── Aba 3: Estatísticas ──────────────────────────────────────────
        ws_stats = wb.create_sheet("Estatísticas")
        stats = self.db.get_stats()
        ws_stats.append(["Métrica", "Valor"])
        for k, v in stats.items():
            ws_stats.append([k, str(v)])

        wb.save(filepath)
        log.info("✔ Excel exportado: %s", filepath)
        return filepath
